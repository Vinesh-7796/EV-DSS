"""Tests for the Phase 2 Document Processing Engine.

Covers PDF, Excel, DBC, Image processors and output validation.
Uses synthetic test files — no real documents required.
"""

import json
import hashlib
from pathlib import Path
from typing import Generator

import pytest

from processing.models.models import (
    Document,
    DocumentMetadata,
    Section,
    Element,
    Chunk,
    Table,
    DBCMessage,
    DBCSignal,
    OCRText,
    BoundingBox,
    Workbook,
    Worksheet,
    Row,
    Cell,
)
from processing.utils.io import save_processed_document, ensure_output_dir
from processing.pdf.processor import PDFProcessor
from processing.excel.processor import ExcelProcessor
from processing.dbc.processor import DBCProcessor
from processing.image.processor import ImageProcessor
from ingestion.models import Document as IngestionDocument


# =========================================================================
# Fixtures
# =========================================================================

@pytest.fixture
def tmp_ingestion_doc(tmp_path: Path) -> IngestionDocument:
    """Helper to create an ingestion document for a file in tmp_path."""
    def _make(filename: str, content: bytes = b"dummy") -> IngestionDocument:
        p = tmp_path / filename
        p.write_bytes(content)
        doc = IngestionDocument(path=p, filename=filename, extension=p.suffix.lower(), size=len(content))
        doc.checksum = hashlib.sha256(content).hexdigest()
        return doc
    return _make


@pytest.fixture
def sample_dbc_text() -> str:
    return '''VERSION "1.0"


NS_ :


BS_:


BU_: VCU MCU BMS


BO_ 385 MotorStatus: 8 MCU
 SG_ MotorRPM : 0|16@1+ (0.125,0) [0|12000] "rpm" VCU
 SG_ MotorTemp : 16|8@1+ (1,-40) [-40|215] "degC" VCU

BO_ 386 HVStatus: 8 BMS
 SG_ BatteryVoltage : 0|16@1+ (0.1,0) [0|500] "V" VCU,MCU
 SG_ MotorCurrent : 16|16@1+ (0.1,-400) [-400|400] "A" MCU

CM_ BO_ 385 "Motor status";
CM_ SG_ 385 MotorRPM "Rotational speed";
CM_ BU_ VCU "Vehicle Control Unit";

BA_ "GenMsgCycleTime" BO_ 385 100;
BA_ "GenMsgCycleTime" BO_ 386 100;
'''


# =========================================================================
# Common Model Tests
# =========================================================================

class TestDataModels:
    """Verify the canonical data models serialize correctly."""

    def test_document_defaults(self) -> None:
        doc = Document()
        assert doc.source == ""
        assert doc.type == ""
        assert doc.metadata.filename == ""
        assert doc.sections == []
        assert doc.tables == []
        assert doc.chunks == []

    def test_document_to_dict(self) -> None:
        doc = Document(
            source="test.pdf",
            type="pdf",
            metadata=DocumentMetadata(filename="test.pdf", page_count=5),
            sections=[Section(id="s1", title="Intro", level=1)],
            tables=[Table(id="t1", headers=["A", "B"], rows=[["1", "2"]])],
            chunks=[Chunk(id="c1", text="hello", page_number=1)],
            raw_text="hello world",
        )
        d = doc.to_dict()
        assert d["source"] == "test.pdf"
        assert d["type"] == "pdf"
        assert d["metadata"]["filename"] == "test.pdf"
        assert d["metadata"]["page_count"] == 5
        assert len(d["sections"]) == 1
        assert len(d["tables"]) == 1
        assert len(d["chunks"]) == 1
        assert d["raw_text"] == "hello world"

    def test_section_hierarchy(self) -> None:
        parent = Section(id="p1", title="Parent", level=1)
        child = Section(id="c1", title="Child", level=2, parent_id="p1")
        parent.subsections.append(child)
        assert parent.title == "Parent"
        assert len(parent.subsections) == 1
        assert parent.subsections[0].id == "c1"

    def test_dbc_message(self) -> None:
        sig = DBCSignal(name="RPM", start_bit=0, length=16, scale=0.125, offset=0, unit="rpm")
        msg = DBCMessage(id=0x181, name="MotorStatus", dlc=8, sender="MCU", signals=[sig])
        assert msg.id == 385
        assert msg.name == "MotorStatus"
        assert msg.signals[0].name == "RPM"

    def test_ocr_text(self) -> None:
        bbox = BoundingBox(x=10, y=20, width=100, height=50)
        ocr = OCRText(text="Hello", confidence=0.95, bbox=bbox)
        assert ocr.text == "Hello"
        assert ocr.bbox.x == 10

    def test_workbook(self) -> None:
        ws = Worksheet(name="Sheet1", rows=[Row(index=1, cells=[Cell(value="A1")])])
        wb = Workbook(filename="test.xlsx", worksheets=[ws])
        assert wb.filename == "test.xlsx"
        assert wb.worksheets[0].name == "Sheet1"

    def test_table_roundtrip(self) -> None:
        t = Table(id="t1", headers=["Col1", "Col2"], rows=[["a", "b"], ["c", "d"]], col_count=2, row_count=2)
        assert t.col_count == 2
        assert t.row_count == 2


# =========================================================================
# PDF Processor Tests
# =========================================================================

class TestPDFProcessor:
    """Verify PDF extraction from a minimal synthetic PDF."""

    def test_processor_properties(self) -> None:
        p = PDFProcessor()
        assert ".pdf" in p.supported_extensions
        assert p.parser_name == "PDFProcessor"

    def test_parse_real_pdf(self, tmp_ingestion_doc) -> None:
        # Create a minimal valid PDF
        minimal_pdf = (
            b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj "
            b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj "
            b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]/Contents 4 0 R"
            b"/Resources<</Font<</F1 5 0 R>>>>>>endobj "
            b"4 0 obj<</Length 44>>stream\nBT /F1 24 Tf 100 700 Td (Hello PDF) Tj ET\nendstream endobj "
            b"5 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj "
            b"xref\n0 6\n0000000000 65535 f \n0000000009 00000 n \n0000000058 00000 n "
            b"\n0000000115 00000 n \n0000000266 00000 n \n0000000361 00000 n "
            b"\ntrailer<</Size 6/Root 1 0 R>>\nstartxref\n442\n%%EOF"
        )
        doc = tmp_ingestion_doc("test.pdf", minimal_pdf)
        processor = PDFProcessor()
        result = processor.parse(doc)
        assert result.processed >= 1
        # Check output file exists
        out = Path("data/processed/pdf/test.json")
        assert out.exists()
        with open(out) as f:
            d = json.load(f)
            assert d["type"] == "pdf"
            assert "Hello PDF" in d["raw_text"]
        out.unlink(missing_ok=True)


# =========================================================================
# Excel Processor Tests
# =========================================================================

class TestExcelProcessor:
    """Verify Excel extraction using openpyxl."""

    def test_processor_properties(self) -> None:
        p = ExcelProcessor()
        assert ".xlsx" in p.supported_extensions
        assert p.parser_name == "ExcelProcessor"

    def test_parse_real_xlsx(self, tmp_path: Path, tmp_ingestion_doc) -> None:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Name"
        ws["B1"] = "Value"
        ws["A2"] = "MotorRPM"
        ws["B2"] = 1200
        ws["A3"] = "BatteryVoltage"
        ws["B3"] = 355
        p = tmp_path / "test.xlsx"
        wb.save(p)
        wb.close()

        doc = IngestionDocument(path=p, filename="test.xlsx", extension=".xlsx", size=p.stat().st_size)
        doc.checksum = hashlib.sha256(p.read_bytes()).hexdigest()
        processor = ExcelProcessor()
        result = processor.parse(doc)
        assert result.processed >= 1

        out = Path("data/processed/xlsx/test.json")
        assert out.exists()
        with open(out) as f:
            d = json.load(f)
            assert d["type"] == "xlsx"
            assert len(d["sections"]) >= 1
            assert d["sections"][0]["title"] == "TestSheet"
        out.unlink(missing_ok=True)


# =========================================================================
# DBC Processor Tests
# =========================================================================

class TestDBCProcessor:
    """Verify DBC file parsing."""

    def test_processor_properties(self) -> None:
        p = DBCProcessor()
        assert ".dbc" in p.supported_extensions
        assert p.parser_name == "DBCProcessor"

    def test_parse_dbc(self, tmp_path: Path, tmp_ingestion_doc, sample_dbc_text) -> None:
        p = tmp_path / "test.dbc"
        p.write_text(sample_dbc_text, encoding="utf-8")

        doc = IngestionDocument(path=p, filename="test.dbc", extension=".dbc", size=p.stat().st_size)
        doc.checksum = hashlib.sha256(sample_dbc_text.encode()).hexdigest()
        processor = DBCProcessor()
        result = processor.parse(doc)
        assert result.processed >= 1

        out = Path("data/processed/dbc/test.json")
        assert out.exists()
        with open(out) as f:
            d = json.load(f)
            assert d["type"] == "dbc"
            assert d["metadata"]["message_count"] == 2
            assert len(d["chunks"]) == 2
        out.unlink(missing_ok=True)

    def test_dbc_nodes(self, sample_dbc_text) -> None:
        processor = DBCProcessor()
        nodes = processor._parse_nodes(sample_dbc_text)
        names = [n.name for n in nodes]
        assert "VCU" in names
        assert "MCU" in names
        assert "BMS" in names

    def test_dbc_messages(self, sample_dbc_text) -> None:
        processor = DBCProcessor()
        messages = processor._parse_messages(sample_dbc_text)
        assert len(messages) == 2
        msg = messages[0]
        assert msg.name == "MotorStatus"
        assert msg.id == 385
        assert msg.dlc == 8
        assert msg.sender == "MCU"
        assert len(msg.signals) == 2
        assert msg.signals[0].name == "MotorRPM"
        assert msg.signals[0].start_bit == 0
        assert msg.signals[0].length == 16
        assert msg.signals[0].scale == 0.125

    def test_dbc_attributes(self, sample_dbc_text) -> None:
        processor = DBCProcessor()
        messages = processor._parse_messages(sample_dbc_text)
        processor._parse_message_attributes(sample_dbc_text, messages)
        msg_map = {m.id: m for m in messages}
        assert msg_map[385].cycle_time == 100.0

    def test_dbc_comments(self, sample_dbc_text) -> None:
        processor = DBCProcessor()
        nodes = processor._parse_nodes(sample_dbc_text)
        messages = processor._parse_messages(sample_dbc_text)
        processor._parse_comments(sample_dbc_text, messages, nodes)
        msg_map = {m.id: m for m in messages}
        assert "Motor status" in msg_map[385].comment
        assert msg_map[385].signals[0].comment == "Rotational speed"
        node_map = {n.name: n for n in nodes}
        assert "Vehicle Control Unit" in node_map["VCU"].comment


# =========================================================================
# Image Processor Tests
# =========================================================================

class TestImageProcessor:
    """Verify image metadata extraction."""

    def test_processor_properties(self) -> None:
        p = ImageProcessor()
        assert ".png" in p.supported_extensions
        assert ".jpg" in p.supported_extensions

    def test_parse_png(self, tmp_path: Path, tmp_ingestion_doc) -> None:
        # Create a valid 1x1 PNG using Pillow
        try:
            from PIL import Image
            import io
            img = Image.new("RGB", (2, 2), color=(255, 0, 0))
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            png_data = buf.getvalue()
        except ImportError:
            pytest.skip("Pillow not installed")

        doc = tmp_ingestion_doc("test.png", png_data)
        processor = ImageProcessor()
        result = processor.parse(doc)
        assert result.processed >= 1

        out = Path("data/processed/png/test.json")
        assert out.exists()
        with open(out) as f:
            d = json.load(f)
            assert d["type"] == "png"
            assert d["metadata"]["image_width"] == 2
            assert d["metadata"]["image_height"] == 2
            assert d["metadata"]["image_format"] == "PNG"
        out.unlink(missing_ok=True)


# =========================================================================
# Output Validation
# =========================================================================

class TestOutputValidation:
    """Verify the structural integrity of processed outputs."""

    def test_all_outputs_have_required_fields(self, tmp_path: Path) -> None:
        """Every processed document must have the mandatory top-level keys."""
        mandatory = {"source", "type", "metadata", "sections", "tables", "images", "chunks", "raw_text"}
        doc = Document(
            source="dummy.pdf",
            type="pdf",
            metadata=DocumentMetadata(filename="dummy.pdf"),
            sections=[Section(id="s1", title="S1")],
        )
        d = doc.to_dict()
        for key in mandatory:
            assert key in d, f"Missing required key: {key}"

    def test_chunks_have_required_fields(self) -> None:
        """Every chunk must carry source context for retrieval."""
        chunk = Chunk(
            id="c1",
            text="some content",
            section_id="s1",
            section_title="Section 1",
            page_number=1,
            source_file="test.pdf",
        )
        assert chunk.id == "c1"
        assert chunk.source_file == "test.pdf"
        assert chunk.page_number == 1

    def test_tables_have_data(self) -> None:
        """Tables must carry headers and row data."""
        t = Table(id="t1", headers=["A", "B"], rows=[["1", "2"], ["3", "4"]], col_count=2, row_count=2)
        assert len(t.rows) == 2
        assert t.rows[0] == ["1", "2"]


# =========================================================================
# Utility Tests
# =========================================================================

class TestProcessingUtils:
    """Verify JSON I/O utilities."""

    def test_ensure_output_dir_creates(self, tmp_path: Path) -> None:
        p = ensure_output_dir("test_type", tmp_path)
        assert p.exists()
        assert p.name == "test_type"

    def test_save_and_load(self, tmp_path: Path) -> None:
        doc = Document(source="test.pdf", type="pdf", metadata=DocumentMetadata(filename="test.pdf"))
        saved = save_processed_document(doc, tmp_path)
        assert saved.exists()
        from processing.utils.io import load_processed_document
        loaded = load_processed_document(saved)
        assert loaded["source"] == "test.pdf"
        assert loaded["type"] == "pdf"
