"""Excel workbook processor.

Extracts worksheets, rows, cells, merged cell ranges, and table
structures from .xlsx files using openpyxl. Produces a standardised
Workbook with typed cell values and header detection.
"""

import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from backend.logger import logger
from ingestion.base.parser import BaseParser
from ingestion.models import Document as IngestionDocument, ProcessingResult, ProcessingStatus
from processing.models.models import (
    ContentNode,
    Document,
    DocumentMetadata,
    Edge,
    IDGenerator,
    ProcessingInfo,
    Reference,
    RelationshipGraph,
    Workbook,
    Worksheet,
    Row,
    Cell,
    Table,
    Section,
    Element,
    Chunk,
    NODE_TYPE_WORKSHEET,
    NODE_TYPE_SPREADSHEET_ROW,
    NODE_TYPE_SPREADSHEET_CELL,
    NODE_TYPE_TABLE,
    NODE_TYPE_PARAGRAPH,
    REF_TYPE_EXCEL,
)
from processing.utils.io import save_processed_document


class ExcelProcessor(BaseParser):
    """Extracts structured content from Excel workbooks."""

    @property
    def supported_extensions(self) -> List[str]:
        return [".xlsx", ".xls"]

    @property
    def parser_name(self) -> str:
        return "ExcelProcessor"

    def parse(self, ingestion_doc: IngestionDocument) -> ProcessingResult:
        logger.info("ExcelProcessor processing: {}", ingestion_doc.filename)
        start = time.time()

        try:
            import openpyxl
        except ImportError:
            ingestion_doc.add_error("openpyxl not installed")
            ingestion_doc.mark(ProcessingStatus.FAILED)
            return ProcessingResult(documents=[ingestion_doc], parser_name=self.parser_name)

        path = ingestion_doc.path
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        except Exception as exc:
            ingestion_doc.add_error(f"Cannot open workbook: {exc}")
            ingestion_doc.mark(ProcessingStatus.FAILED)
            return ProcessingResult(documents=[ingestion_doc], parser_name=self.parser_name)

        worksheets: List[Worksheet] = []
        all_tables: List[Table] = []
        all_raw_text: List[str] = []
        table_counter = 0

        for ws_index, ws_name in enumerate(wb.sheetnames, start=1):
            ws = wb[ws_name]
            sheet_rows: List[Row] = []
            merged = list(ws.merged_cells.ranges)
            merged_set: set = set()
            for mr in merged:
                for row in range(mr.min_row, mr.max_row + 1):
                    for col in range(mr.min_col, mr.max_col + 1):
                        merged_set.add((row, col))

            # Detect header row (first row typically)
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0

            for r in range(1, max_row + 1):
                cells: List[Cell] = []
                is_header_row = (r == 1)
                for c in range(1, max_col + 1):
                    raw_val = ws.cell(row=r, column=c).value
                    val = self._clean_cell(raw_val)
                    col_letter = openpyxl.utils.get_column_letter(c)
                    cell = Cell(
                        value=val,
                        column=col_letter,
                        row_index=r,
                        column_index=c,
                        is_merged=(r, c) in merged_set,
                        is_header=is_header_row and bool(val),
                    )
                    cells.append(cell)
                    if val is not None and str(val).strip():
                        all_raw_text.append(f"[{ws_name}] {col_letter}{r}: {val}")

                sheet_rows.append(Row(index=r, cells=cells, is_header=is_header_row))

            # Build tables from cell data
            if sheet_rows:
                table_counter += 1
                headers = [c.value if isinstance(c.value, str) else str(c.value or "")
                           for c in sheet_rows[0].cells] if sheet_rows else []
                data_rows = []
                for row in sheet_rows[1:]:
                    data_rows.append([str(c.value or "") for c in row.cells])
                tbl = Table(
                    id=f"tbl_{table_counter}",
                    caption=f"Worksheet: {ws_name}",
                    headers=headers,
                    rows=data_rows,
                    col_count=max_col,
                    row_count=len(data_rows),
                )
                all_tables.append(tbl)

            ws_model = Worksheet(
                name=ws_name,
                rows=sheet_rows,
                merged_cells=[(mr.min_row, mr.min_col, mr.max_row, mr.max_col) for mr in merged],
                column_count=max_col,
                row_count=max_row,
            )
            worksheets.append(ws_model)

        wb.close()

        # ── Build output ──
        meta = DocumentMetadata(
            filename=ingestion_doc.filename,
            file_size=ingestion_doc.size,
            checksum=ingestion_doc.checksum,
            sheet_count=len(worksheets),
            processed_at=time.strftime("%Y-%m-%dT%H:%M:%S"),
            processing_time_s=time.time() - start,
        )

        workbook = Workbook(
            filename=ingestion_doc.filename,
            metadata=meta,
            worksheets=worksheets,
        )

        # Flatten to generic Document for unified storage
        raw_text = "\n".join(all_raw_text)
        sections = [
            Section(
                id=f"sheet_{i}",
                title=ws.name,
                level=1,
                elements=[
                    Element(type="paragraph", content=f"Sheet: {ws.name} ({ws.row_count} rows x {ws.column_count} cols)")
                ],
            )
            for i, ws in enumerate(worksheets)
        ]
        chunks = [
            Chunk(
                id=f"chunk_sheet_{i}",
                text=f"Worksheet '{ws.name}':\n" + "\n".join(
                    str(c.value or "") for r in ws.rows for c in r.cells if c.value
                ),
                section_id=f"sheet_{i}",
                section_title=ws.name,
                source_file=ingestion_doc.filename,
            )
            for i, ws in enumerate(worksheets)
        ]

        ext = Path(ingestion_doc.filename).suffix.lstrip(".").lower() or "excel"
        output = Document(
            source=ingestion_doc.filename,
            type=ext,
            metadata=meta,
            sections=sections,
            tables=all_tables,
            chunks=chunks,
            raw_text=raw_text,
        )

        # ── Build CDS ──
        content_nodes, graph = self._build_cds(worksheets, all_tables, ingestion_doc.filename)
        output.content_nodes = content_nodes
        output.relationship_graph = graph
        output.processing_info = ProcessingInfo(
            processor_name=self.parser_name,
            processing_time_s=time.time() - start,
            processed_at=time.strftime("%Y-%m-%dT%H:%M:%S"),
            schema_version="2.0",
        )

        save_processed_document(output)
        ingestion_doc.mark(ProcessingStatus.COMPLETED)
        logger.info("ExcelProcessor done: {} sheets, {} tables, {} chunks",
                     meta.sheet_count, len(all_tables), len(chunks))
        return ProcessingResult(documents=[ingestion_doc], parser_name=self.parser_name,
                                duration_s=time.time() - start)

    # ────────────────────────────────────────────
    #  CDS Builder
    # ────────────────────────────────────────────

    @staticmethod
    def _build_cds(
        worksheets: List[Worksheet],
        all_tables: List[Table],
        filename: str,
    ) -> Tuple[List[ContentNode], RelationshipGraph]:
        """Build CDS ContentNode hierarchy from Excel worksheets."""
        id_gen = IDGenerator(doc_number=1)
        all_nodes: Dict[str, ContentNode] = {}
        edges: List[Edge] = []
        root_nodes: List[ContentNode] = []

        for ws_idx, ws in enumerate(worksheets):
            ws_node_id = id_gen.worksheet_id(None)
            ws_node = ContentNode(
                id=ws_node_id,
                type=NODE_TYPE_WORKSHEET,
                content=ws.name,
                reference=Reference(
                    type=REF_TYPE_EXCEL,
                    location={"worksheet": ws.name},
                ),
            )
            all_nodes[ws_node_id] = ws_node
            edges.append(Edge(source=ws_node_id, target=id_gen.doc_id,
                              relationship_type="child_of"))

            for row in ws.rows:
                row_node_id = id_gen.row_id(ws_node_id)
                row_node = ContentNode(
                    id=row_node_id,
                    type=NODE_TYPE_SPREADSHEET_ROW,
                    content={"index": row.index, "is_header": row.is_header},
                    reference=Reference(
                        type=REF_TYPE_EXCEL,
                        location={"worksheet": ws.name, "row": row.index},
                    ),
                    parent_id=ws_node_id,
                )
                all_nodes[row_node_id] = row_node
                edges.append(Edge(source=row_node_id, target=ws_node_id,
                                  relationship_type="child_of"))

                for cell in row.cells:
                    cell_node_id = id_gen.cell_id(row_node_id)
                    col_ref = f"{cell.column}{cell.row_index}"
                    cell_node = ContentNode(
                        id=cell_node_id,
                        type=NODE_TYPE_SPREADSHEET_CELL,
                        content=cell.value,
                        reference=Reference(
                            type=REF_TYPE_EXCEL,
                            location={
                                "worksheet": ws.name,
                                "row": cell.row_index,
                                "column": cell.column,
                                "cell": col_ref,
                            },
                        ),
                        parent_id=row_node_id,
                        metadata={
                            "is_merged": cell.is_merged,
                            "is_header": cell.is_header,
                            "column_index": cell.column_index,
                        },
                    )
                    all_nodes[cell_node_id] = cell_node
                    row_node.children.append(cell_node)

                ws_node.children.append(row_node)

            root_nodes.append(ws_node)

        # Table nodes (in addition to worksheet rows)
        for tbl in all_tables:
            tbl_node_id = id_gen.table_id(None)
            tbl_node = ContentNode(
                id=tbl_node_id,
                type=NODE_TYPE_TABLE,
                content={"headers": tbl.headers, "caption": tbl.caption},
                reference=Reference(
                    type=REF_TYPE_EXCEL,
                    location={"worksheet": tbl.caption.replace("Worksheet: ", ""), "table": tbl.id},
                ),
            )
            all_nodes[tbl_node_id] = tbl_node
            edges.append(Edge(source=tbl_node_id, target=id_gen.doc_id,
                              relationship_type="child_of"))
            root_nodes.append(tbl_node)

        return root_nodes, RelationshipGraph(nodes=all_nodes, edges=edges)

    @staticmethod
    def _clean_cell(val: Any) -> Any:
        """Normalise common Excel value types."""
        if val is None:
            return ""
        if isinstance(val, float) and val == int(val):
            return int(val)
        return val
